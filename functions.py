# functions.py

import openpyxl  # Biblioteka do obsługi plików Excel
from openpyxl.styles import PatternFill, Font  # Style do formatowania komórek
from tasks import TodoTask, InProgressTask, DoneTask  # Import klas zadań

# Stała z nazwą pliku Excela
NAZWA_PLIKU = "Tasks.xlsx"


# Funkcja sprawdzająca, czy plik istnieje, a jeśli nie, tworząca nowy plik
def sprawdz_lub_utworz_plik():
    try:
        # Próba załadowania istniejącego pliku
        skoroszyt = openpyxl.load_workbook(NAZWA_PLIKU)
    except FileNotFoundError:
        # Jeśli plik nie istnieje, tworzymy nowy skoroszyt
        skoroszyt = openpyxl.Workbook()
        arkusz = skoroszyt.active
        arkusz.title = "Tasks"  # Ustawienie nazwy arkusza
        arkusz.append(["Title", "Status"])  # Dodanie nagłówków kolumn
        skoroszyt.save(NAZWA_PLIKU)  # Zapisanie nowego pliku
    return skoroszyt

# Funkcja ładująca zadania z pliku Excel
def zaladuj_zadania():
    # Załadowanie lub utworzenie pliku
    skoroszyt = sprawdz_lub_utworz_plik()
    arkusz = skoroszyt.active
    zadania = []
    # Iterowanie przez wiersze w arkuszu, pomijając nagłówki
    for wiersz in arkusz.iter_rows(min_row=2, values_only=True):
        tytul, status = wiersz
        # Tworzenie odpowiednich obiektów zadań na podstawie statusu
        if status == "Do zrobienia":
            zadania.append(TodoTask(tytul))
        elif status == "W trakcie":
            zadania.append(InProgressTask(tytul))
        elif status == "Wykonane":
            zadania.append(DoneTask(tytul))
    return zadania

# Funkcja zapisująca zadania do pliku Excel
def zapisz_zadania(zadania):
    # Tworzenie nowego skoroszytu
    skoroszyt = openpyxl.Workbook()
    arkusz = skoroszyt.active
    arkusz.title = "Tasks"  # Ustawienie nazwy arkusza
    arkusz.append(["Title", "Status"])  # Dodanie nagłówków kolumn
    # Ustawienie rozmiaru kolumn
    arkusz.column_dimensions['A'].width = 5  # Ustawiamy szerokość kolumny A na 5
    arkusz.column_dimensions['B'].width = 12  # Ustawiamy szerokość kolumny B na 20

    # Iterowanie przez zadania i zapisywanie ich do arkusza
    for zadanie in zadania:
        arkusz.append([zadanie.tytul, zadanie.status])
        # Wybór koloru wypełnienia komórek w zależności od statusu zadania
        if zadanie.status == "Do zrobienia":
            wypelnienie = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        elif zadanie.status == "W trakcie":
            wypelnienie = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            czcionka = Font(bold=True)
        elif zadanie.status == "Wykonane":
            wypelnienie = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        # Zastosowanie stylu do komórek
        for komorka in arkusz.iter_rows(min_row=arkusz.max_row, max_row=arkusz.max_row, min_col=1, max_col=2):
            for k in komorka:
                k.fill = wypelnienie
                if zadanie.status == "W trakcie":
                    k.font = czcionka
    skoroszyt.save(NAZWA_PLIKU)  # Zapisanie pliku


# Funkcja dodająca nowe zadanie
def dodaj_zadanie(tytul):
    zadania = zaladuj_zadania()  # Załadowanie istniejących zadań
    istniejące_tytuly = [zadanie.tytul for zadanie in zadania]
    # Sprawdzenie, czy zadanie o podanym tytule już istnieje
    if tytul in istniejące_tytuly:
        licznik = 1
        nowy_tytul = f"{tytul}-{licznik}"
        # Jeśli tytuł istnieje, tworzymy nowy, unikalny tytuł
        while nowy_tytul in istniejące_tytuly:
            licznik += 1
            nowy_tytul = f"{tytul}-{licznik}"
        print(f"Zadanie o nazwie '{tytul}' już istnieje. Nowe zadanie będzie miało nazwę '{nowy_tytul}'.")
        tytul = nowy_tytul

    zadania.append(TodoTask(tytul))  # Dodanie nowego zadania do listy
    zapisz_zadania(zadania)  # Zapisanie zaktualizowanej listy zadań

# Funkcja kończąca zadanie
def zakoncz_zadanie():
    zadania = zaladuj_zadania()  # Załadowanie istniejących zadań
    # Szukanie zadania w trakcie i pytanie użytkownika, czy chce je zakończyć
    for zadanie in zadania:
        if zadanie.status == "W trakcie":
            decyzja = input(f"Istnieje zadanie w trakcie: '{zadanie.tytul}'. Czy chcesz je zakończyć? (t/n): ").lower()
            if decyzja == 't':
                print("Zakończono!")
                zadanie.status = "Wykonane"
                zapisz_zadania(zadania)  # Zapisanie zaktualizowanej listy zadań
                zasugeruj_nastepne_zadanie(zadania)  # Sugerowanie następnego zadania do rozpoczęcia
                return
            elif decyzja == 'n':
                return  # Zakończenie funkcji bez zakończenia zadania
            elif decyzja != 'n' and decyzja != 't':
                print("Nieprawidłowa odpowiedź. Proszę odpowiedzieć 't' lub 'n'.")
                break
    print("Nie znaleziono zadania w trakcie do zakończenia.")

# Funkcja sugerująca następne zadanie do rozpoczęcia
def zasugeruj_nastepne_zadanie(zadania):
    # Szukanie zadań o statusie "Do zrobienia"
    zadania_do_zrobienia = [zadanie for zadanie in zadania if zadanie.status == "Do zrobienia"]

    if not zadania_do_zrobienia:
        print("Brak zadań do zrobienia.")
        return

    print("Dostępne zadania do rozpoczęcia:")
    for zadanie in zadania_do_zrobienia:
        print(f"- {zadanie.tytul}")

    while True:
        wybor = input("Wpisz nazwę zadania, które chcesz rozpocząć (pozostaw puste aby anulować): ")
        if wybor == "":
            print("Anulowano wybór zadania.")
            return
        # Sprawdzenie, czy wybrane zadanie znajduje się na liście dostępnych zadań
        if any(zadanie.tytul == wybor for zadanie in zadania_do_zrobienia):
            rozpocznij_zadanie(wybor)
            return
        else:
            print("Nie znaleziono zadania o podanej nazwie. Spróbuj ponownie.")


# Funkcja wyświetlająca zadania
def wyswietl_zadania():
    zadania = zaladuj_zadania()  # Załadowanie istniejących zadań
    # Iterowanie przez zadania i wyświetlanie ich z odpowiednim formatowaniem
    for zadanie in zadania:
        if zadanie.status == "Do zrobienia":
            print(f"{zadanie}")
        elif zadanie.status == "W trakcie":
            print(f"{zadanie}")
        elif zadanie.status == "Wykonane":
            print(f"{zadanie}")

def menu():
    # Pętla nieskończona, która pozwala na wybór opcji z menu
    while True:
        # Wyświetlanie menu
        print("\nMenedżer Zadań")
        print("1. Dodaj nowe zadanie")
        print("2. Rozpocznij zadanie")
        print("3. Zakończ zadanie")
        print("4. Wyświetl zadania")
        print("5. Wyjście")

        # Pobranie wyboru użytkownika
        wybor = input("Wybierz opcję: ")

        # Sprawdzenie, którą opcję wybrał użytkownik
        if wybor == '1':
            # Jeśli wybrał opcję 1, pytamy o tytuł zadania i dodajemy nowe zadanie
            tytul = input("Podaj tytuł zadania: ")
            dodaj_zadanie(tytul)
        elif wybor == '2':
            # Jeśli wybrał opcję 2, pytamy o tytuł zadania do rozpoczęcia
            tytul = input("Podaj tytuł zadania do rozpoczęcia: ")
            rozpocznij_zadanie(tytul)
        elif wybor == '3':
            zakoncz_zadanie()
        elif wybor == '4':
            # Jeśli wybrał opcję 4, wyświetlamy wszystkie zadania
            wyswietl_zadania()
        elif wybor == '5':
            # Jeśli wybrał opcję 5, kończymy pętlę i program
            break
        else:
            # Jeśli użytkownik podał nieprawidłową opcję, informujemy go o tym
            print("Nieprawidłowa opcja, spróbuj ponownie.")